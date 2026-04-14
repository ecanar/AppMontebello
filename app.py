from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import cast, Integer
from datetime import datetime, timezone, timedelta
import os
import json
from collections import defaultdict
import calendar
from io import BytesIO
from dotenv import load_dotenv
from openai import OpenAI

TZ_COLOMBIA = timezone(timedelta(hours=-5))

# Cargar variables de entorno
load_dotenv()

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'default_secret_key')

login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Debes iniciar sesión para acceder a esta página.'

# Configuración de la base de datos (Soporta Railway y Local)
database_uri = os.getenv('DATABASE_URL', os.getenv('SQLALCHEMY_DATABASE_URI', 'sqlite:///montebello.db'))
# Railway usa "postgres://" pero SQLAlchemy requiere "postgresql://"
if database_uri.startswith("postgres://"):
    database_uri = database_uri.replace("postgres://", "postgresql://", 1)
# Si la URI no empieza con un esquema válido, usar SQLite como fallback
if not database_uri.startswith(("postgresql://", "postgresql+", "sqlite://", "mysql://")):
    print(f"ADVERTENCIA: DATABASE_URL inválida ('{database_uri[:30]}...'), usando SQLite.")
    database_uri = 'sqlite:///montebello.db'

# Intentar psycopg3 si psycopg2 no está disponible
if database_uri.startswith("postgresql://"):
    try:
        import psycopg2
    except ImportError:
        try:
            import psycopg
            database_uri = database_uri.replace("postgresql://", "postgresql+psycopg://", 1)
        except ImportError:
            pass

app.config['SQLALCHEMY_DATABASE_URI'] = database_uri
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {'pool_pre_ping': True}

db = SQLAlchemy(app)

# Modelo de Usuario
class Usuario(db.Model, UserMixin):
    __tablename__ = 'usuarios'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    activo = db.Column(db.Boolean, default=True)
    es_admin = db.Column(db.Boolean, default=False)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    @property
    def is_active(self):
        return self.activo

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.es_admin:
            flash('Acceso restringido a administradores.')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated

@login_manager.user_loader
def load_user(user_id):
    return Usuario.query.get(int(user_id))

# Tabla de Productos
class Producto(db.Model):
    __tablename__ = 'productos'
    id_Prod = db.Column(db.Integer, primary_key=True)
    Nom_Prod = db.Column(db.String(100), nullable=False)
    Medida = db.Column(db.String(10), nullable=False)  # sac, lbs, ata, rac, cart, uni, cub
    Id_Prov = db.Column(db.Integer, db.ForeignKey('proveedores.Id_Prov'), nullable=False)
    
    # Relación con proveedor
    proveedor = db.relationship('Proveedor', backref=db.backref('productos', lazy=True))

# Tabla de Proveedores
class Proveedor(db.Model):
    __tablename__ = 'proveedores'
    Id_Prov = db.Column(db.Integer, primary_key=True)
    Nom_Prov = db.Column(db.String(100), nullable=False)
    Num_Ced = db.Column(db.String(20), nullable=False)
    Num_Anden = db.Column(db.String(10))
    Num_Puesto = db.Column(db.String(10))
    Alias = db.Column(db.String(100))

# Tabla Compras del Día
class CompraDia(db.Model):
    __tablename__ = 'compras_dia'
    Id_Lin_Comp = db.Column(db.Integer, primary_key=True)
    Id_Comp = db.Column(db.Integer, nullable=False)
    Fec_Comp = db.Column(db.Date, nullable=False, default=lambda: datetime.now(TZ_COLOMBIA).date())
    Id_Prod = db.Column(db.Integer, db.ForeignKey('productos.id_Prod'), nullable=False)
    Cant_Ped = db.Column(db.Float, nullable=False)
    Cant_Bod = db.Column(db.Float, nullable=False)
    Cant_Comp = db.Column(db.Float, nullable=False)
    Val_Pag = db.Column(db.Float, nullable=False)
    Id_Prov = db.Column(db.Integer, db.ForeignKey('proveedores.Id_Prov'), nullable=False)
    
    # Relaciones
    producto = db.relationship('Producto', backref=db.backref('compras_dia', lazy=True))
    proveedor_compra = db.relationship('Proveedor', backref=db.backref('compras_dia', lazy=True))

# Tabla Histórico de Compras
class HistoricoCompra(db.Model):
    __tablename__ = 'historico_compras'
    Id_Lin_Comp = db.Column(db.Integer, primary_key=True)
    Id_Comp = db.Column(db.Integer, nullable=False)
    Fec_Comp = db.Column(db.Date, nullable=False)
    Id_Prod = db.Column(db.Integer, db.ForeignKey('productos.id_Prod'), nullable=False)
    Cant_Ped = db.Column(db.Float, nullable=False)
    Cant_Comp = db.Column(db.Float, nullable=False)
    Cant_Bod = db.Column(db.Float, nullable=False)
    Val_Pag = db.Column(db.Float, nullable=False)
    Id_Prov = db.Column(db.Integer, db.ForeignKey('proveedores.Id_Prov'), nullable=False)
    
    # Relaciones
    producto_h = db.relationship('Producto', backref=db.backref('historico_compras', lazy=True))
    proveedor_h = db.relationship('Proveedor', backref=db.backref('historico_compras', lazy=True))

# Tabla Pedidos de Compra
class PedidoCompra(db.Model):
    __tablename__ = 'pedidos_compra'
    Id_Lin_Ped = db.Column(db.Integer, primary_key=True)
    Id_Lista = db.Column(db.Integer, nullable=False)  # ID secuencial anual
    Id_Prod = db.Column(db.Integer, db.ForeignKey('productos.id_Prod'), nullable=False)
    Cant_Ped = db.Column(db.Float, nullable=False)
    Cant_Bod = db.Column(db.Float, nullable=False)
    Fec_Ped = db.Column(db.Date, nullable=False, default=lambda: datetime.now(TZ_COLOMBIA).date())
    
    # Relaciones
    producto_pedido = db.relationship('Producto', backref=db.backref('pedidos_compra', lazy=True))

# Tabla de Medidas
class Medida(db.Model):
    __tablename__ = 'medidas'
    id_Medida = db.Column(db.Integer, primary_key=True)
    Cod_Medida = db.Column(db.String(10), nullable=False, unique=True)
    Nom_Medida = db.Column(db.String(50), nullable=False)

# Rutas de autenticación
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        usuario = Usuario.query.filter_by(username=username).first()
        if usuario and usuario.check_password(password) and usuario.activo:
            login_user(usuario)
            next_page = request.args.get('next')
            return redirect(next_page or url_for('index'))
        flash('Usuario o contraseña incorrectos.')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# Rutas principales
@app.route('/')
@login_required
def index():
    return render_template('index.html')

# CRUD Productos
@app.route('/productos')
@login_required
def productos():
    productos = Producto.query.order_by(Producto.Nom_Prod).all()
    proveedores = Proveedor.query.order_by(Proveedor.Nom_Prov).all()
    medidas = Medida.query.order_by(Medida.Nom_Medida).all()
    return render_template('productos.html', productos=productos, proveedores=proveedores, medidas=medidas)

@app.route('/productos/add', methods=['POST'])
@login_required
def add_producto():
    if request.method == 'POST':
        nombre = request.form.get('nombre')
        medida = request.form.get('medida')
        id_prov = request.form.get('proveedor')
        
        if not nombre or not medida or not id_prov:
            flash('Por favor complete todos los campos obligatorios.')
            return redirect(url_for('productos'))
        
        try:
            nuevo_producto = Producto(Nom_Prod=nombre, Medida=medida, Id_Prov=id_prov)
            db.session.add(nuevo_producto)
            db.session.commit()
            flash('Producto agregado exitosamente!')
        except Exception as e:
            db.session.rollback()
            flash(f'Error al agregar producto: {str(e)}')
            
        return redirect(url_for('productos'))

@app.route('/productos/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_producto(id):
    producto = Producto.query.get_or_404(id)
    proveedores = Proveedor.query.order_by(Proveedor.Nom_Prov).all()
    medidas = Medida.query.order_by(Medida.Nom_Medida).all()
    
    if request.method == 'POST':
        nombre = request.form.get('nombre')
        medida = request.form.get('medida')
        id_prov = request.form.get('proveedor')
        
        if not nombre or not medida or not id_prov:
            flash('Por favor complete todos los campos obligatorios.')
            return render_template('edit_producto.html', producto=producto, proveedores=proveedores, medidas=medidas)
        
        try:
            producto.Nom_Prod = nombre
            producto.Medida = medida
            producto.Id_Prov = id_prov
            db.session.commit()
            flash('Producto actualizado exitosamente!')
            return redirect(url_for('productos'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar producto: {str(e)}')
    
    return render_template('edit_producto.html', producto=producto, proveedores=proveedores, medidas=medidas)

@app.route('/productos/delete/<int:id>')
@login_required
def delete_producto(id):
    producto = Producto.query.get_or_404(id)
    try:
        db.session.delete(producto)
        db.session.commit()
        flash('Producto eliminado exitosamente!')
    except Exception as e:
        db.session.rollback()
        flash('Error al eliminar: puede que el producto esté referenciado en compras o pedidos.')
    return redirect(url_for('productos'))

# CRUD Proveedores
@app.route('/proveedores')
@login_required
def proveedores():
    proveedores = Proveedor.query.order_by(Proveedor.Nom_Prov).all()
    return render_template('proveedores.html', proveedores=proveedores)

@app.route('/proveedores/add', methods=['POST'])
@login_required
def add_proveedor():
    if request.method == 'POST':
        nombre = request.form.get('nombre')
        cedula = request.form.get('cedula')
        anden = request.form.get('anden')
        puesto = request.form.get('puesto')
        
        if not nombre or not cedula:
            flash('Nombre y Cédula son campos obligatorios.')
            return redirect(url_for('proveedores'))

        if Proveedor.query.filter(Proveedor.Nom_Prov.ilike(nombre)).first():
            flash(f'El proveedor "{nombre}" ya existe.')
            return redirect(url_for('proveedores'))
            
        try:
            alias = request.form.get('alias', '').strip() or None
            nuevo_proveedor = Proveedor(Nom_Prov=nombre, Num_Ced=cedula, Num_Anden=anden, Num_Puesto=puesto, Alias=alias)
            db.session.add(nuevo_proveedor)
            db.session.commit()
            flash('Proveedor agregado exitosamente!')
        except Exception as e:
            db.session.rollback()
            flash(f'Error al agregar proveedor: {str(e)}')
            
        return redirect(url_for('proveedores'))

@app.route('/proveedores/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_proveedor(id):
    proveedor = Proveedor.query.get_or_404(id)
    
    if request.method == 'POST':
        nombre = request.form.get('nombre')
        cedula = request.form.get('cedula')
        anden = request.form.get('anden')
        puesto = request.form.get('puesto')
        
        if not nombre or not cedula:
            flash('Nombre y Cédula son campos obligatorios.')
            return render_template('edit_proveedor.html', proveedor=proveedor)
            
        try:
            proveedor.Nom_Prov = nombre
            proveedor.Num_Ced = cedula
            proveedor.Num_Anden = anden
            proveedor.Num_Puesto = puesto
            proveedor.Alias = request.form.get('alias', '').strip() or None
            db.session.commit()
            flash('Proveedor actualizado exitosamente!')
            return redirect(url_for('proveedores'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar proveedor: {str(e)}')
    
    return render_template('edit_proveedor.html', proveedor=proveedor)

@app.route('/proveedores/delete/<int:id>')
@login_required
def delete_proveedor(id):
    proveedor = Proveedor.query.get_or_404(id)
    try:
        db.session.delete(proveedor)
        db.session.commit()
        flash('Proveedor eliminado exitosamente!')
    except Exception as e:
        db.session.rollback()
        flash('Error al eliminar: puede que el proveedor tenga productos asociados.')
    return redirect(url_for('proveedores'))

# Compras del Día
@app.route('/compras')
@login_required
def compras():
    compras = CompraDia.query.join(Proveedor, CompraDia.Id_Prov == Proveedor.Id_Prov).order_by(cast(Proveedor.Num_Anden, Integer), cast(Proveedor.Num_Puesto, Integer)).all()
    productos = Producto.query.order_by(Producto.Nom_Prod).all()
    proveedores = Proveedor.query.order_by(Proveedor.Nom_Prov).all()
    hoy = datetime.now(TZ_COLOMBIA).date()
    primera = CompraDia.query.first()
    fec_compras = primera.Fec_Comp if primera else hoy
    return render_template('compras.html', compras=compras, productos=productos, proveedores=proveedores,
                           fec_compras=fec_compras, hoy=hoy)

@app.route('/compras/add', methods=['POST'])
@login_required
def add_compra():
    if request.method == 'POST':
        try:
            # Obtener el siguiente Id_Comp
            ultima_compra = CompraDia.query.order_by(CompraDia.Id_Comp.desc()).first()
            id_comp = (ultima_compra.Id_Comp + 1) if ultima_compra else 1
            
            id_prod = request.form.get('producto')
            cant_ped = float(request.form.get('cant_ped', 0))
            cant_bod = float(request.form.get('cant_bod', 0))
            cant_comp = float(request.form.get('cant_comp', 0))
            val_pag = float(request.form.get('val_pag', 0))
            
            if not id_prod:
                flash('Debe seleccionar un producto.')
                return redirect(url_for('compras'))
            
            # Obtener el proveedor del producto
            producto = Producto.query.get(id_prod)
            if not producto:
                flash('Producto no encontrado.')
                return redirect(url_for('compras'))
            
            id_prov = producto.Id_Prov
            
            nueva_compra = CompraDia(
                Id_Comp=id_comp,
                Id_Prod=id_prod,
                Cant_Ped=cant_ped,
                Cant_Bod=cant_bod,
                Cant_Comp=cant_comp,
                Val_Pag=val_pag,
                Id_Prov=id_prov
            )
            
            db.session.add(nueva_compra)
            db.session.commit()
            flash('Compra agregada exitosamente!')
        except Exception as e:
            db.session.rollback()
            flash(f'Error al agregar compra: {str(e)}')
            
        return redirect(url_for('compras'))

@app.route('/compras/update/<int:id>', methods=['POST'])
@login_required
def update_compra(id):
    compra = CompraDia.query.get_or_404(id)
    try:
        data = request.get_json()
        if 'cant_comp' in data:
            compra.Cant_Comp = float(data['cant_comp'])
        if 'val_pag' in data:
            compra.Val_Pag = float(data['val_pag'])
        db.session.commit()
        total = sum(c.Val_Pag for c in CompraDia.query.all())
        return {'ok': True, 'total': round(total, 2)}
    except Exception as e:
        db.session.rollback()
        return {'ok': False, 'error': str(e)}, 400

@app.route('/compras/delete/<int:id>')
@login_required
def delete_compra(id):
    compra = CompraDia.query.get_or_404(id)
    try:
        db.session.delete(compra)
        db.session.commit()
        flash('Compra eliminada exitosamente!')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar: {str(e)}')
    return redirect(url_for('compras'))

# CRUD Medidas
@app.route('/medidas')
@login_required
def medidas():
    medidas = Medida.query.order_by(Medida.Nom_Medida).all()
    return render_template('medidas.html', medidas=medidas)

@app.route('/medidas/add', methods=['POST'])
@login_required
def add_medida():
    cod = request.form.get('cod_medida', '').strip().lower()
    nom = request.form.get('nom_medida', '').strip()
    if not cod or not nom:
        flash('Complete todos los campos.')
        return redirect(url_for('medidas'))
    if Medida.query.filter_by(Cod_Medida=cod).first():
        flash(f'El código "{cod}" ya existe. Use otro código o edite la medida existente.')
        return redirect(url_for('medidas'))
    try:
        db.session.add(Medida(Cod_Medida=cod, Nom_Medida=nom))
        db.session.commit()
        flash('Medida agregada exitosamente!')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}')
    return redirect(url_for('medidas'))

@app.route('/medidas/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_medida(id):
    medida = Medida.query.get_or_404(id)
    if request.method == 'POST':
        cod = request.form.get('cod_medida', '').strip().lower()
        nom = request.form.get('nom_medida', '').strip()
        if not cod or not nom:
            flash('Complete todos los campos.')
            return render_template('edit_medida.html', medida=medida)
        try:
            medida.Cod_Medida = cod
            medida.Nom_Medida = nom
            db.session.commit()
            flash('Medida actualizada!')
            return redirect(url_for('medidas'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}')
    return render_template('edit_medida.html', medida=medida)

@app.route('/medidas/delete/<int:id>')
@login_required
def delete_medida(id):
    medida = Medida.query.get_or_404(id)
    try:
        db.session.delete(medida)
        db.session.commit()
        flash('Medida eliminada!')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar: {str(e)}')
    return redirect(url_for('medidas'))

# Histórico de Compras
@app.route('/historico')
@login_required
def historico():
    historico = HistoricoCompra.query.order_by(HistoricoCompra.Id_Comp.desc(), HistoricoCompra.Id_Lin_Comp).all()
    return render_template('historico.html', historico=historico)

@app.route('/historico/plantilla')
@login_required
def historico_plantilla():
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Compras'
    headers = ['Fecha (YYYY-MM-DD)', 'Producto', 'Proveedor', 'Cant_Ped', 'Cant_Comp', 'Cant_Bod', 'Val_Pag']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    ws2 = wb.create_sheet('Productos Válidos')
    ws2.append(['Nombre Producto', 'Medida'])
    for p in Producto.query.order_by(Producto.Nom_Prod).all():
        ws2.append([p.Nom_Prod, p.Medida])
    ws3 = wb.create_sheet('Proveedores Válidos')
    ws3.append(['Nombre Proveedor'])
    for p in Proveedor.query.order_by(Proveedor.Nom_Prov).all():
        ws3.append([p.Nom_Prov])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='plantilla_historico.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/historico/importar', methods=['POST'])
@login_required
def historico_importar():
    import openpyxl
    if 'archivo' not in request.files:
        flash('No se seleccionó archivo.')
        return redirect(url_for('historico'))
    archivo = request.files['archivo']
    if not archivo.filename.endswith('.xlsx'):
        flash('Solo se aceptan archivos .xlsx')
        return redirect(url_for('historico'))
    try:
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        ultimo_hist = HistoricoCompra.query.order_by(HistoricoCompra.Id_Comp.desc()).first()
        ultimo_dia  = CompraDia.query.order_by(CompraDia.Id_Comp.desc()).first()
        base_id = max(
            ultimo_hist.Id_Comp if ultimo_hist else 0,
            ultimo_dia.Id_Comp  if ultimo_dia  else 0
        )
        fecha_to_id = {}
        errores = []
        importados = 0
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            fecha_raw, nom_prod, nom_prov, cant_ped, cant_comp, cant_bod, val_pag = (list(row) + [None]*7)[:7]
            if isinstance(fecha_raw, datetime):
                fecha = fecha_raw.date()
            else:
                try:
                    fecha = datetime.strptime(str(fecha_raw).strip(), '%Y-%m-%d').date()
                except Exception:
                    errores.append(f'Fila {i}: fecha inválida "{fecha_raw}"')
                    continue
            if fecha not in fecha_to_id:
                base_id += 1
                fecha_to_id[fecha] = base_id
            id_comp = fecha_to_id[fecha]
            producto = Producto.query.filter(Producto.Nom_Prod.ilike(str(nom_prod or '').strip())).first()
            if not producto:
                errores.append(f'Fila {i}: producto "{nom_prod}" no encontrado')
                continue
            proveedor = Proveedor.query.filter(Proveedor.Nom_Prov.ilike(str(nom_prov or '').strip())).first()
            if not proveedor:
                errores.append(f'Fila {i}: proveedor "{nom_prov}" no encontrado')
                continue
            db.session.add(HistoricoCompra(
                Id_Comp=id_comp, Fec_Comp=fecha, Id_Prod=producto.id_Prod,
                Cant_Ped=float(cant_ped or 0), Cant_Comp=float(cant_comp or 0),
                Cant_Bod=float(cant_bod or 0), Val_Pag=float(val_pag or 0),
                Id_Prov=proveedor.Id_Prov
            ))
            importados += 1
        db.session.commit()
        msg = f'{importados} registros importados.'
        if errores:
            msg += ' Errores: ' + '; '.join(errores[:5])
            if len(errores) > 5:
                msg += f' ... y {len(errores)-5} más.'
        flash(msg)
    except Exception as e:
        db.session.rollback()
        flash(f'Error al importar: {str(e)}')
    return redirect(url_for('historico'))

# Mover compras del día a histórico
@app.route('/mover_historico', methods=['POST'])
@login_required
def mover_historico():
    try:
        compras_hoy = CompraDia.query.all()
        if not compras_hoy:
            flash('No hay compras para mover al histórico.')
            return redirect(url_for('compras'))

        fecha_str = request.form.get('fecha_real', '')
        try:
            fecha_real = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except ValueError:
            fecha_real = datetime.now(TZ_COLOMBIA).date()

        for compra in compras_hoy:
            historico_registro = HistoricoCompra(
                Id_Comp=compra.Id_Comp,
                Fec_Comp=fecha_real,
                Id_Prod=compra.Id_Prod,
                Cant_Ped=compra.Cant_Ped,
                Cant_Comp=compra.Cant_Comp,
                Cant_Bod=compra.Cant_Bod,
                Val_Pag=compra.Val_Pag,
                Id_Prov=compra.Id_Prov
            )
            db.session.add(historico_registro)

        CompraDia.query.delete()
        db.session.commit()
        flash('Compras movidas al histórico exitosamente!')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al mover al histórico: {str(e)}')

    return redirect(url_for('compras'))

# Pedidos de Compra
@app.route('/pedidos')
@login_required
def pedidos():
    pedidos = PedidoCompra.query.join(Producto, PedidoCompra.Id_Prod == Producto.id_Prod).order_by(Producto.Nom_Prod).all()
    productos = Producto.query.order_by(Producto.Nom_Prod).all()
    primera_compra = CompraDia.query.first()
    hay_compras = primera_compra is not None
    id_comp_actual = primera_compra.Id_Comp if primera_compra else None
    n_compras_actual = CompraDia.query.count()
    return render_template('pedidos.html', pedidos=pedidos, productos=productos,
                           hay_compras=hay_compras, id_comp_actual=id_comp_actual,
                           n_compras_actual=n_compras_actual)

@app.route('/pedidos/add', methods=['POST'])
@login_required
def add_pedido():
    if request.method == 'POST':
        try:
            # Obtener el siguiente Id_Lista secuencial para el año actual
            año_actual = datetime.now().year
            ultimo_pedido = PedidoCompra.query.filter(
                db.extract('year', PedidoCompra.Fec_Ped) == año_actual
            ).order_by(PedidoCompra.Id_Lista.desc()).first()
            
            id_lista = (ultimo_pedido.Id_Lista + 1) if ultimo_pedido else 1
            
            id_prod = request.form.get('producto')
            # Si el campo está vacío, se asigna 0.0 automáticamente
            try:
                cant_ped = float(request.form.get('cant_ped') or 0)
                cant_bod = float(request.form.get('cant_bod') or 0)
            except ValueError:
                flash('Las cantidades deben ser valores numéricos.')
                return redirect(url_for('pedidos'))
            
            if not id_prod:
                flash('Debe seleccionar un producto.')
                return redirect(url_for('pedidos'))
                
            nuevo_pedido = PedidoCompra(
                Id_Lista=id_lista,
                Id_Prod=id_prod,
                Cant_Ped=cant_ped,
                Cant_Bod=cant_bod
            )
            
            db.session.add(nuevo_pedido)
            db.session.commit()
            flash('Pedido agregado exitosamente!')
        except Exception as e:
            db.session.rollback()
            flash(f'Error al agregar pedido: {str(e)}')
            
        return redirect(url_for('pedidos'))

@app.route('/pedidos/delete/<int:id>')
@login_required
def delete_pedido(id):
    pedido = PedidoCompra.query.get_or_404(id)
    try:
        db.session.delete(pedido)
        db.session.commit()
        flash('Pedido eliminado exitosamente!')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar: {str(e)}')
    return redirect(url_for('pedidos'))

@app.route('/transferir_pedidos')
@login_required
def transferir_pedidos():
    try:
        pedidos_pendientes = PedidoCompra.query.all()
        if not pedidos_pendientes:
            flash('No hay pedidos pendientes para transferir.')
            return redirect(url_for('compras'))

        nueva = request.args.get('nueva')   # None | '0' | '1'
        borrar = request.args.get('borrar', '0') == '1'
        hoy = datetime.now(TZ_COLOMBIA).date()
        hay_compras = CompraDia.query.count() > 0

        if hay_compras and nueva is None:
            flash('Hay ítems en la lista de Compras. Confirme la acción desde la página de Pedidos.')
            return redirect(url_for('pedidos'))

        if hay_compras and nueva == '0':
            id_comp = CompraDia.query.first().Id_Comp
        else:
            ultima_dia  = CompraDia.query.order_by(CompraDia.Id_Comp.desc()).first()
            ultimo_hist = HistoricoCompra.query.order_by(HistoricoCompra.Id_Comp.desc()).first()
            id_comp = max(
                ultima_dia.Id_Comp  if ultima_dia  else 0,
                ultimo_hist.Id_Comp if ultimo_hist else 0
            ) + 1
            if hay_compras and borrar:
                CompraDia.query.delete()
                db.session.flush()

        transferidos = 0
        for pedido in pedidos_pendientes:
            producto = Producto.query.get(pedido.Id_Prod)
            if not producto:
                continue
            nueva_compra = CompraDia(
                Id_Comp=id_comp,
                Fec_Comp=hoy,
                Id_Prod=pedido.Id_Prod,
                Cant_Ped=pedido.Cant_Ped,
                Cant_Bod=pedido.Cant_Bod,
                Cant_Comp=0,
                Val_Pag=0,
                Id_Prov=producto.Id_Prov
            )
            db.session.add(nueva_compra)
            transferidos += 1

        PedidoCompra.query.delete()
        db.session.commit()
        flash(f'Se han transferido {transferidos} pedidos a la Compra Nº {id_comp}.')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al transferir pedidos: {str(e)}')

    return redirect(url_for('compras'))

# Cambiar propia contraseña
@app.route('/mi-cuenta', methods=['GET', 'POST'])
@login_required
def mi_cuenta():
    if request.method == 'POST':
        actual = request.form.get('password_actual', '')
        nueva = request.form.get('password_nueva', '')
        confirmar = request.form.get('password_confirmar', '')
        if not current_user.check_password(actual):
            flash('La contraseña actual es incorrecta.')
        elif not nueva:
            flash('La nueva contraseña no puede estar vacía.')
        elif nueva != confirmar:
            flash('La nueva contraseña y la confirmación no coinciden.')
        else:
            current_user.set_password(nueva)
            db.session.commit()
            flash('Contraseña actualizada exitosamente!')
            return redirect(url_for('index'))
    return render_template('mi_cuenta.html')

# Gestión de Usuarios
@app.route('/usuarios')
@admin_required
def usuarios():
    usuarios = Usuario.query.order_by(Usuario.username).all()
    return render_template('usuarios.html', usuarios=usuarios)

@app.route('/usuarios/add', methods=['POST'])
@admin_required
def add_usuario():
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '')
    if not username or not password:
        flash('Usuario y contraseña son obligatorios.')
        return redirect(url_for('usuarios'))
    if Usuario.query.filter_by(username=username).first():
        flash(f'El usuario "{username}" ya existe.')
        return redirect(url_for('usuarios'))
    try:
        nuevo = Usuario(username=username)
        nuevo.set_password(password)
        db.session.add(nuevo)
        db.session.commit()
        flash(f'Usuario "{username}" creado exitosamente!')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al crear usuario: {str(e)}')
    return redirect(url_for('usuarios'))

@app.route('/usuarios/cambiar_password/<int:id>', methods=['POST'])
@admin_required
def cambiar_password(id):
    usuario = Usuario.query.get_or_404(id)
    password = request.form.get('password', '')
    if not password:
        flash('La contraseña no puede estar vacía.')
        return redirect(url_for('usuarios'))
    try:
        usuario.set_password(password)
        db.session.commit()
        flash(f'Contraseña de "{usuario.username}" actualizada!')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}')
    return redirect(url_for('usuarios'))

@app.route('/usuarios/toggle/<int:id>')
@admin_required
def toggle_usuario(id):
    usuario = Usuario.query.get_or_404(id)
    if usuario.username == 'admin' and current_user.id == usuario.id:
        flash('No podés desactivar tu propio usuario admin.')
        return redirect(url_for('usuarios'))
    try:
        usuario.activo = not usuario.activo
        db.session.commit()
        estado = 'activado' if usuario.activo else 'desactivado'
        flash(f'Usuario "{usuario.username}" {estado}.')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}')
    return redirect(url_for('usuarios'))

@app.route('/usuarios/toggle_admin/<int:id>')
@admin_required
def toggle_admin(id):
    usuario = Usuario.query.get_or_404(id)
    if usuario.id == current_user.id:
        flash('No podés cambiar tus propios permisos de admin.')
        return redirect(url_for('usuarios'))
    try:
        usuario.es_admin = not usuario.es_admin
        db.session.commit()
        estado = 'otorgado' if usuario.es_admin else 'quitado'
        flash(f'Permiso admin {estado} a "{usuario.username}".')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}')
    return redirect(url_for('usuarios'))

@app.route('/usuarios/delete/<int:id>')
@admin_required
def delete_usuario(id):
    usuario = Usuario.query.get_or_404(id)
    if usuario.id == current_user.id:
        flash('No podés eliminar tu propio usuario.')
        return redirect(url_for('usuarios'))
    try:
        db.session.delete(usuario)
        db.session.commit()
        flash(f'Usuario "{usuario.username}" eliminado.')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}')
    return redirect(url_for('usuarios'))

with app.app_context():
    try:
        # Migración previa: agregar es_admin ANTES de que el ORM intente usarla
        try:
            from sqlalchemy import text
            with db.engine.connect() as conn:
                conn.execute(text('ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS es_admin BOOLEAN DEFAULT FALSE'))
                conn.commit()
        except Exception:
            pass  # La tabla aún no existe; db.create_all() la crea con la columna
        try:
            with db.engine.connect() as conn:
                conn.execute(text('ALTER TABLE proveedores ADD COLUMN IF NOT EXISTS "Alias" VARCHAR(100)'))
                conn.commit()
        except Exception:
            pass
        db.create_all()
        admin_pwd = os.getenv('ADMIN_PASSWORD', 'admin123')
        admin = Usuario.query.filter_by(username='admin').first()
        if not admin:
            admin = Usuario(username='admin', es_admin=True)
            admin.set_password(admin_pwd)
            db.session.add(admin)
            db.session.commit()
            print('Usuario admin creado.')
        else:
            changed = False
            if not admin.es_admin:
                admin.es_admin = True
                changed = True
            if os.getenv('ADMIN_PASSWORD'):
                admin.set_password(admin_pwd)
                changed = True
                print('Contraseña admin reseteada desde ADMIN_PASSWORD.')
            if changed:
                db.session.commit()
    except Exception as e:
        print(f'Advertencia al crear tablas: {e}')


@app.route('/consultas-ia', methods=['GET', 'POST'])
@login_required
def consultas_ia():
    respuesta = None
    error = None
    pregunta = None
    if request.method == 'POST':
        pregunta = request.form.get('pregunta', '').strip()
        try:
            api_key = os.getenv('OPENROUTER_API_KEY')
            if not api_key:
                error = 'No se encontró la API Key de OpenRouter.'
            else:
                client = OpenAI(
                    base_url='https://openrouter.ai/api/v1',
                    api_key=api_key
                )

                historico = HistoricoCompra.query.order_by(HistoricoCompra.Id_Comp.desc()).limit(500).all()
                compras_hoy = CompraDia.query.all()
                proveedores_list = Proveedor.query.order_by(Proveedor.Nom_Prov).all()
                productos_list = Producto.query.order_by(Producto.Nom_Prod).all()
                pedidos_list = PedidoCompra.query.order_by(PedidoCompra.Fec_Ped.desc()).all()

                contexto = "Lista de proveedores registrados:\n"
                for p in proveedores_list:
                    contexto += f"- {p.Nom_Prov}, Andén={p.Num_Anden}, Puesto={p.Num_Puesto}\n"
                contexto += "\nLista de productos registrados:\n"
                for p in productos_list:
                    contexto += f"- {p.Nom_Prod} ({p.Medida}), proveedor={p.proveedor.Nom_Prov}\n"
                contexto += "\nPedidos de compra pendientes:\n"
                for r in pedidos_list:
                    contexto += f"- Fecha={r.Fec_Ped}, Producto={r.producto_pedido.Nom_Prod}, cant_pedida={r.Cant_Ped}, cant_bodega={r.Cant_Bod}\n"
                contexto += "\nCompras del día actual:\n"
                for r in compras_hoy:
                    contexto += f"- Comp={r.Id_Comp}, Fecha={r.Fec_Comp}, Producto={r.producto.Nom_Prod} ({r.producto.Medida}), cant_pedida={r.Cant_Ped}, cant_comprada={r.Cant_Comp}, valor=${r.Val_Pag}, proveedor={r.proveedor_compra.Nom_Prov}\n"
                contexto += "\nHistórico de todas las compras:\n"
                for r in historico:
                    contexto += f"- Comp={r.Id_Comp}, Fecha={r.Fec_Comp}, Producto={r.producto_h.Nom_Prod} ({r.producto_h.Medida}), cant_pedida={r.Cant_Ped}, cant_comprada={r.Cant_Comp}, valor=${r.Val_Pag}, proveedor={r.proveedor_h.Nom_Prov}\n"

                sistema = (
                    "Eres un asistente de una app de compras de mercado. "
                    "Responde SIEMPRE en español, de forma clara y concisa. "
                    "El campo 'Id_Comp' o 'Comp' es el número de compra que agrupa todos los productos comprados en una misma salida al mercado. "
                    "Nunca mezcles otros idiomas. No incluyas explicaciones técnicas ni notas al pie."
                )
                prompt = f"{contexto}\nPregunta: {pregunta}"
                response = client.chat.completions.create(
                    model='openai/gpt-4o-mini',
                    messages=[
                        {'role': 'system', 'content': sistema},
                        {'role': 'user', 'content': prompt}
                    ]
                )
                respuesta = response.choices[0].message.content
        except Exception as e:
            error = f'Error al consultar IA: {str(e)}'
    return render_template('consultas_ia.html', respuesta=respuesta, error=error, pregunta=pregunta)

@app.route('/analisis')
@login_required
def analisis():
    MESES_ES = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']

    fec_ini_str = request.args.get('fec_ini', '').strip()
    fec_fin_str = request.args.get('fec_fin', '').strip()
    mes_str     = request.args.get('mes', '').strip()
    modo = request.args.get('modo', 'compra')
    if modo not in ('compra', 'mes'):
        modo = 'compra'

    fec_ini = fec_fin = None
    if modo == 'compra':
        try:
            fec_ini = datetime.strptime(fec_ini_str, '%Y-%m-%d').date() if fec_ini_str else None
        except ValueError:
            fec_ini = None; fec_ini_str = ''
        try:
            fec_fin = datetime.strptime(fec_fin_str, '%Y-%m-%d').date() if fec_fin_str else None
        except ValueError:
            fec_fin = None; fec_fin_str = ''
    elif modo == 'mes' and mes_str:
        try:
            yr, mn = int(mes_str[:4]), int(mes_str[5:7])
            fec_ini = datetime(yr, mn, 1).date()
            fec_fin = datetime(yr, mn, calendar.monthrange(yr, mn)[1]).date()
        except (ValueError, IndexError):
            mes_str = ''

    todos = HistoricoCompra.query.order_by(HistoricoCompra.Fec_Comp).all()
    registros = [r for r in todos
                 if (fec_ini is None or r.Fec_Comp >= fec_ini)
                 and (fec_fin is None or r.Fec_Comp <= fec_fin)]

    extra = dict(active_mode=modo, fec_ini_val=fec_ini_str, fec_fin_val=fec_fin_str, mes_val=mes_str)
    empty = dict(sin_datos=True, total_gasto=0, n_semanas=0, n_productos=0,
                 n_proveedores=0, n_compras=0,
                 data_compra='{}', data_semana='{}', data_mes='{}',
                 top_productos='[]', top_frec='[]', top_proveedores='[]',
                 bod_top='[]', precio_evolucion_prods='[]', **extra)
    if not registros:
        return render_template('analisis.html', **empty)

    def _agrupar(regs, key_fn, label_fn, top5):
        gasto_g   = defaultdict(float)
        ped_g     = defaultdict(float)
        comp_g    = defaultdict(float)
        salidas_g = defaultdict(set)
        precio_g  = {}
        keys_set  = set()
        for r in regs:
            k  = key_fn(r)
            np = r.producto_h.Nom_Prod
            keys_set.add(k)
            gasto_g[k]  += r.Val_Pag
            ped_g[k]    += r.Cant_Ped
            comp_g[k]   += r.Cant_Comp
            salidas_g[k].add(r.Id_Comp)
            if r.Cant_Comp > 0:
                precio_g.setdefault(np, {}).setdefault(k, [0.0, 0.0])
                precio_g[np][k][0] += r.Val_Pag
                precio_g[np][k][1] += r.Cant_Comp
        keys    = sorted(keys_set)
        labels  = [label_fn(k) for k in keys]
        gasto   = [round(gasto_g[k], 2) for k in keys]
        cumpl   = [round(comp_g[k]/ped_g[k]*100, 1) if ped_g[k] > 0 else 0.0 for k in keys]
        salidas = [len(salidas_g[k]) for k in keys]
        precio_evol = {}
        for prod in top5:
            vals = []
            for k in keys:
                d = precio_g.get(prod, {}).get(k)
                vals.append(round(d[0]/d[1], 2) if d and d[1] > 0 else None)
            precio_evol[prod] = vals
        return {'labels': labels, 'gasto': gasto, 'cumpl': cumpl,
                'salidas': salidas, 'precio': precio_evol}

    gasto_producto  = defaultdict(float)
    gasto_prov      = defaultdict(float)
    bod_producto    = defaultdict(list)
    frec_producto   = defaultdict(int)
    semanas_set     = set()

    for r in registros:
        np = r.producto_h.Nom_Prod
        semanas_set.add(tuple(r.Fec_Comp.isocalendar()[:2]))
        gasto_producto[np]                   += r.Val_Pag
        gasto_prov[r.proveedor_h.Nom_Prov]  += r.Val_Pag
        bod_producto[np].append(r.Cant_Bod)
        frec_producto[np]                    += 1

    top_productos   = sorted(gasto_producto.items(), key=lambda x: x[1], reverse=True)[:10]
    top_frec        = sorted(frec_producto.items(),  key=lambda x: x[1], reverse=True)[:10]
    top_proveedores = sorted(gasto_prov.items(),     key=lambda x: x[1], reverse=True)
    top5_prods      = [p[0] for p in top_productos[:5]]
    bod_top         = sorted(
        {p: round(sum(v)/len(v), 2) for p, v in bod_producto.items() if v}.items(),
        key=lambda x: x[1], reverse=True)[:10]

    data_compra = _agrupar(registros,
        key_fn=lambda r: r.Id_Comp,
        label_fn=lambda k: f"C{k}",
        top5=top5_prods)

    data_semana = _agrupar(registros,
        key_fn=lambda r: tuple(r.Fec_Comp.isocalendar()[:2]),
        label_fn=lambda k: f"S{k[1]:02d}/{str(k[0])[2:]}",
        top5=top5_prods)

    data_mes = _agrupar(registros,
        key_fn=lambda r: (r.Fec_Comp.year, r.Fec_Comp.month),
        label_fn=lambda k: f"{MESES_ES[k[1]-1]}/{str(k[0])[2:]}",
        top5=top5_prods)

    total_gasto = round(sum(gasto_producto.values()), 2)
    n_compras   = len(set(r.Id_Comp for r in registros))

    return render_template('analisis.html',
        sin_datos=False,
        total_gasto=total_gasto,
        n_semanas=len(semanas_set),
        n_productos=len(gasto_producto),
        n_proveedores=len(gasto_prov),
        n_compras=n_compras,
        data_compra=json.dumps(data_compra),
        data_semana=json.dumps(data_semana),
        data_mes=json.dumps(data_mes),
        top_productos=json.dumps([[p[0], round(p[1],2)] for p in top_productos]),
        top_frec=json.dumps([[p[0], p[1]] for p in top_frec]),
        top_proveedores=json.dumps([[p[0], round(p[1],2)] for p in top_proveedores]),
        bod_top=json.dumps([[p[0], p[1]] for p in bod_top]),
        precio_evolucion_prods=json.dumps(top5_prods),
        active_mode=modo,
        fec_ini_val=fec_ini_str,
        fec_fin_val=fec_fin_str,
        mes_val=mes_str,
    )


if __name__ == '__main__':
    # Railway asigna el puerto automáticamente en la variable de entorno PORT
    port = int(os.environ.get("PORT", 5000)) 
    app.run(host='0.0.0.0', port=port)

